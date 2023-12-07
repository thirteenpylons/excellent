import contextlib
import re
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime, timedelta
from typing import List, Dict, Set, Tuple

from mlib.key_data import tech_info


class ExcelProcessor:
    def __init__(self, file_path: str):
        self.file_path: str = file_path
        self.workbook: Workbook = load_workbook(file_path)
        self.worksheet: Worksheet = self.workbook.active

    def auto_size_columns(self, worksheet: Worksheet) -> None:
        for column in worksheet.columns:
            max_length = 0
            for cell in column:
                with contextlib.suppress(TypeError):
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    def extract_dates_from_range(self, cell_range: str) -> Set[str]:
        cell_value = self.worksheet[cell_range].value
        # Extract the start and end dates from the cell value
        # This assumes that the cell value is a string like "Report Period: 11/28/2023 - 12/28/2023"
        date_matches = re.findall(r'\b\d{1,2}/\d{1,2}/\d{4}\b', cell_value)
        if len(date_matches) != 2:
            raise ValueError(f"Expected to find two dates in the cell value but found {len(date_matches)}")
        start_date_str, end_date_str = date_matches
        start_date = datetime.strptime(start_date_str, '%m/%d/%Y').date()
        end_date = datetime.strptime(end_date_str, '%m/%d/%Y').date()

        # Generate all dates in the range
        current_date = start_date
        dates = set()
        while current_date <= end_date:
            dates.add(current_date.strftime('%m/%d/%Y'))
            current_date += timedelta(days=1)  # Increment the day

        return dates
            
    def create_headers(self, dates: Set[str]) -> List[str]:
        headers: List[str] = ['Name', 'Tech #']
        for date in sorted(dates):
            headers.extend([f"{date} Morning", f"{date} Afternoon"])
        return headers

    def extract_eng_circuit(self, job_comment: str) -> Tuple[str, str]:
        # Regex pattern to extract ENG and circuit type
        eng_pattern = re.compile(r'ENG-\d+')
        circuit_pattern = re.compile(r'\b(FIA|DIA|DEDICATED INTERNET SERVICE|CARRIER E-ACCESS|HVOF|HV|HVOD|SBB|BENCH TEST|FC\+|TRUNK|MNS|MRS|MNE|MANAGED NETWORK EDGE|AGG SWITCH)\b')
        action_pattern = re.compile(r'\b(MW|INSTALL|PRE|EQUIPMENT PU|EQUIPMENT P/U|EQUIP PU|EQUIP P/U|SWEEP)\b')

        # Search for ENG and circuit type in the job comment
        eng_match = eng_pattern.search(job_comment)
        circuit_match = circuit_pattern.search(job_comment)
        action_match = action_pattern.search(job_comment)

        # Extract ENG and circuit type if found
        eng = eng_match[0] if eng_match else ''
        circuit_type = circuit_match[0] if circuit_match else ''
        action = action_match[0] if action_match else ''

        return eng, action, circuit_type

    def process_rows(self) -> None:
        dates: Set[str] = self.extract_dates_from_range('A7')  # Only once
        headers: List[str] = self.create_headers(dates)  # Only once

        updated_wb: Workbook = Workbook()
        updated_ws: Worksheet = updated_wb.active
        updated_ws.append(headers)

        # Create a dictionary to hold all job details for each technician and date
        tech_jobs: Dict[str, Dict[str, List[str]]] = {}
        for row in self.worksheet.iter_rows(min_row=11, values_only=True):
            tech_num: str = str(row[2])  # Tech number in column 'C'
            if tech_num not in tech_jobs:
                tech_jobs[tech_num] = {header: '' for header in headers[2:]}  # Skip Name and Tech #
            
            tech_name: str = tech_info.get(tech_num, 'Unknown Tech')
            date_str: str = row[5].strftime('%m/%d/%Y') if isinstance(row[5], datetime) else row[5]
            timeslot: str = row[6]  # Timeslot in column 'G'
            job_comment: str = row[7]  # Job comment in column 'H'
            
            eng, action, circuit_type = self.extract_eng_circuit(job_comment)
            job_details: str = f"{eng}\r\n{action} {circuit_type} {row[4]}\r\n {timeslot}" # writing WO data to cell, row[4] is the address

            time_category: str = "Afternoon" if "PM" in timeslot.upper() else "Morning"
            date_header: str = f"{date_str} {time_category}"

            if date_header in tech_jobs[tech_num]:
                tech_jobs[tech_num][date_header] += f"{job_details}"  # Append job details to the existing string
            else:
                tech_jobs[tech_num][date_header] = job_details

        # Write the combined job details for each technician to the worksheet
        for tech_num, jobs in tech_jobs.items():
            tech_name: str = tech_info.get(tech_num, 'Unknown Tech')
            new_row: List[str] = [tech_name, tech_num] + [jobs[header] for header in headers[2:]]
            updated_ws.append(new_row)

        self.auto_size_columns(updated_ws)  # Auto-size columns after filling data
        updated_wb.save('updated_calendar.xlsx')  # Saving the file


if __name__ == "__main__":
    original_file_path: str = './data/Copy of Enterprise WO 30 Day.xlsx'
    processor: ExcelProcessor = ExcelProcessor(original_file_path)
    processor.process_rows()  # Process rows and save within the method